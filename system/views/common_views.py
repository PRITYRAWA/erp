from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from ..serializers.common_serializers import *
from ..models.columns import Column
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.decorators import action
from system import utils
import openpyxl
from django.db.models import Q
from openpyxl_image_loader import SheetImageLoader
from system.models.dataset import Table, Data

# Generate Token Manually
def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }

#**********************************Function To Extract Data From Excel File******************************************#
def extracting_data(file):
    xl_data = []
    heading = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    max_col = sheet.max_column
    for i in range(1, max_col+1):
        cell_obj = sheet.cell(row = 1, column = i)
        heading.append((cell_obj.value).lower())
    for row in sheet.iter_rows(min_row=2):
        data_dict={}
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        for i in range(len(row_data)):
            data_dict[heading[i]]=row_data[i]
        xl_data.append(data_dict)
    return(xl_data)

class ButtonViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Buttons to be modified.
    """
    queryset = Button.objects.all()
    serializer_class = ButtonSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
class CurrencyViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Currency to be modified.
    """
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    try:
                        serializer=CurrencySerializer(data=data_dict[i], context={'request':request})
                        if serializer.is_valid(raise_exception=True):
                            serializer.save()
                            count += 1
                    except Exception as e:
                        print("CURRENCY ERROR >>>> ",str(e)) 
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
           
class TagViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Tag to be modified.
    """
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
class LanguageViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Language to be modified.
    """
    queryset = Language.objects.all()
    serializer_class = LanguageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
class CountryViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Country to be modified.
    """
    queryset = Country.objects.all()
    serializer_class = CountrySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            symbol_selector = Selectors.objects.get(selector__icontains='country_symbol_position')
            money_selector = Selectors.objects.get(selector__icontains='country_money_format')
            date_selector = Selectors.objects.get(selector__icontains='country_date_format')
            time_selector = Selectors.objects.get(selector__icontains='country_time_format')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    print(data_dict[i])
                    currency=data_dict[i].get('currency')
                    country=data_dict[i].get('country')
                    symbol_position = data_dict[i].pop('currency_symbol_position')
                    money_format = data_dict[i].pop('money_format')
                    data_format = data_dict[i].pop('date_format')
                    time_format = data_dict[i].pop('time_format')
                    if currency and country:
                        try:
                            currency_rec = Currency.objects.get(Q(name=currency)| Q(code=currency))
                            symbol_rec = Choice.objects.filter(choice_name=symbol_position, selector=symbol_selector.id)
                            money_rec = Choice.objects.filter(choice_name=money_format, selector=money_selector.id)
                            date_rec = Choice.objects.filter(choice_name=data_format, selector=date_selector.id)
                            time_rec = Choice.objects.filter(choice_name=time_format, selector=time_selector.id)

                            data_dict[i]['currency']=currency_rec.id
                            if symbol_rec:
                                data_dict[i]['symbol_position']=symbol_rec.values()[0]['id']
                            if money_rec:
                                data_dict[i]['money_format']=money_rec.values()[0]['id']
                            if date_rec:
                                data_dict[i]['date_format']=date_rec.values()[0]['id']
                            if time_rec:
                                data_dict[i]['time_format']=time_rec.values()[0]['id']

                            serializer=CountrySerializer(data=data_dict[i], context={'request':request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count += 1
                        except Exception as e:
                            #print("COUNTRY ERROR >>>> ", str(e), data_dict[i])
                            pass
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
        
class StateViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows State to be modified.
    """
    queryset = State.objects.all()
    serializer_class = StateSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            print("demo")
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                print("file >> ", file)
                count = 0
                defective_data=[]
                for i in range(len(data_dict)):
                    country=data_dict[i].get('country')
                    state_name = data_dict[i].get('name')
                    if country and state_name:
                        try:
                            country_rec = Country.objects.get(Q(country__iname=country) | Q(country=country))
                            data_dict[i]['country']=country_rec.id
                            
                            state_check=State.objects.filter(name=state_name, country=data_dict[i]['country'])
                            if not state_check:
                                serializer=StateSerializer(data=data_dict[i], context={'request':request})
                                if serializer.is_valid(raise_exception=True):
                                    serializer.save()
                                    count += 1
                        except Exception as e:
                            defective_data.append(country)
                            print(str(e))
                            pass
                if defective_data:
                    defective_data = {
                        "inavlid_country" : f"These {set(defective_data)} are the invalid country names."
                    }
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
    
class StageViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Stage to be modified.
    """
    queryset = Stage.objects.all()
    serializer_class = StageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    def list(self, request, args, *kwargs):
        queryset = Stage.objects.filter().all()
        serializers = StageSerializer(queryset, many = True, context = {"request": request})
        return super().list(request, args, *kwargs)
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                defective_data=[]
                for i in range(len(data_dict)):
                    if data_dict[i] != None:
                        stage_name = data_dict[i]['stage']
                        if stage_name:
                            form_name = data_dict[i]['form']
                            form_rec = Form.objects.filter(form=form_name)
                            if form_rec:
                                req_action = data_dict[i].pop('required action buttons')
                                opt_action = data_dict[i].pop('optional action buttons')
                                data_dict[i]['form']=form_rec.values()[0]['id']
                                stage_rec = Stage.objects.filter(stage = stage_name,form_id=form_rec.values()[0]['id'])
                                if stage_rec:
                                    stage_id = stage_rec.values()[0]['id']
                                else:
                                    serializer=StageSerializer(data=data_dict[i],context={'request': request})
                                    if serializer.is_valid():
                                        serializer.save()
                                        count += 1
                                    stage_rec = Stage.objects.get(stage = stage_name,form_id=form_rec.values()[0]['id'])
                                    stage_id=stage_rec.id
                                if req_action:
                                    text_split = req_action.split(',')
                                    for i in range (len(text_split)):
                                        check = StageAction.objects.filter(stage_id=stage_id,action=text_split[i])
                                        if not check:
                                            StageAction.objects.create(stage_id=stage_id,action=text_split[i],required=True)
                                if opt_action:
                                    text_split = opt_action.split(',')
                                    for i in range (len(text_split)):
                                        check = StageAction.objects.filter(stage_id=stage_id,action=text_split[i])
                                        if not check:
                                            StageAction.objects.create(stage_id=stage_id,action=text_split[i],optional=True)
                            else:
                                defective_data.append(form_name)
                if defective_data:
                    defective_data = {
                        "missing_form" : f"These {set(defective_data)} are the invalid form names."
                    }
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))


class StageActionViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Configuration to be modified.
    """
    queryset = StageAction.objects.all()
    serializer_class = StageActionSerializer       
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

 
class ConfigurationViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Configuration to be modified.
    """
    queryset = Configuration.objects.all()
    serializer_class = ConfigurationSerializer       
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    conf_name = data_dict[i]["configuration"]
                    editable = data_dict[i]["editable"]
                    if conf_name:
                        conf = Configuration.objects.filter(configuration=conf_name)
                        if not conf:
                            if editable == 'yes' or 'Yes':
                                data_dict[i]["editable"]=True
                            else:
                                data_dict[i]["editable"] = False
                            serializer=ConfigurationSerializer(data=data_dict[i],context={'request': request})
                            if serializer.is_valid():
                                serializer.save()
                                count += 1
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
            
class TerritoriesViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Territories to be modified.
    """
    queryset = Territories.objects.all()
    serializer_class = TerritoriesSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class SelectorViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Selector to be modified.
    """
    queryset = Selectors.objects.all()
    serializer_class = SelectorSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    selector = data_dict[i].get('selector')
                    if selector:
                        try:
                            serializer=SelectorSerializer(data=data_dict[i], context={'request':request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count += 1
                        except Exception as e:
                            pass          
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
            
class ChoiceViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Choice to be modified.
    """
    queryset = Choice.objects.all()
    serializer_class = ChoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            # have_trans = False
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    # if 'label (us english)' in data_dict[i]:
                    #     label=data_dict[i].pop('label (us english)')
                    #     have_trans = True
                    selector = data_dict[i]['selector']
                    choice = data_dict[i]["choice_name"]
                    default = data_dict[i]["default"]
                    if selector:
                        selector = utils.encode_api_name(selector)
                        check = Selectors.objects.filter(selector=selector)
                        if check:
                            data_dict[i]['selector'] = check.values()[0]['id']
                            selector = check.values()[0]['id']
                        else:
                            new_selector = Selectors.objects.create(selector=selector, type='user')
                            selector = new_selector.id
                            data_dict[i]['selector'] = new_selector.id
                        if choice:
                            choice_name=utils.encode_api_name(data_dict[i]["choice_name"])
                            choice_rec = Choice.objects.filter(choice_name=choice_name, selector = selector)
                            if not choice_rec:
                                if default == 'yes' or 'Yes':
                                    data_dict[i]['default'] = True
                                else:
                                    data_dict[i]['default'] = False
                                try:
                                    serializer = ChoiceSerializer(data=data_dict[i], context={'request':request})
                                    if serializer.is_valid():
                                        serializer.save()
                                        count += 1
                                    # choice_id = Choice.objects.get(id=serializer.data.get('id'))
                                    # if have_trans == True:
                                    #     language = get_current_user_language(request.user)
                                    #     lang = Language.objects.get(name=language)
                                    #     if label:
                                    #         check=Translation.objects.filter(label=label,language_id=lang.id)
                                    #         if not check:
                                    #             new_label = Translation.objects.create(label=label,language_id=lang.id)
                                    #     label_rec = Translation.objects.get(label=label,language_id=lang.id)
                                    #     trans = TranslationChoice.objects.filter(choice_id=choice_id, translation_id=label_rec.id)
                                    #     if not trans:
                                    #         TranslationChoice.objects.create(choice_id=choice_id, translation_id=label_rec.id)
                                except Exception as e:
                                    print(str(e))
                                    pass
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class FormViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Form to be modified.
    """
    queryset = Form.objects.all()
    serializer_class = FormSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    if data_dict[i] != None:
                        form_name = data_dict[i]['form']
                        form_rec = Form.objects.filter(form=form_name)
                        if not form_rec:
                            serializer=FormSerializer(data=data_dict[i],context={'request': request})
                            if serializer.is_valid():
                                serializer.save()
                                count += 1
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class ListViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows List to be modified.
    """
    queryset = List.objects.all()
    serializer_class = ListSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    # Related List 
    @action(detail=True, methods=['get'], url_path = "columns")
    def get_columns(self, request, pk=None): 
        queryset = Column.objects.filter(list = pk) 
        serializer = ColumnsSerializer(queryset, many = True, context={'request': request})         
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    list_data = data_dict[i]['system_name']
                    category=data_dict[i].pop('category')
                    label = data_dict[i].pop('label (us english)')
                    data_source = data_dict[i]['primary_table']
                    list_type = data_dict[i]['list_type']
                    visibility = data_dict[i]['visibility']
                    table_rec = Table.objects.filter(table=data_source)
                    list_type_rec = Choice.objects.filter(choice_name=list_type)
                    visibility_rec = Choice.objects.filter(choice_name=visibility)
                    if list_data and table_rec and list_type_rec and visibility_rec:
                        list_rec = List.objects.filter(system_name=list_data)
                        if list_rec:
                            list_id=list_rec.values()[0]['id']
                        else:
                            data_dict[i]['primary_table']=table_rec.values()[0]['id']
                            data_dict[i]['list_type']=list_type_rec.values()[0]['id']
                            data_dict[i]['visibility']=visibility_rec.values()[0]['id']
                            list_serializers = ListSerializer(data=data_dict[i], context={'request': request})
                            if list_serializers.is_valid(raise_exception=True):
                                list_serializers.save()
                                count = count + 1    
                            list_id=list_serializers.data.get('id')
                        language = get_current_user_language(request.user)
                        lang = Language.objects.get(name=language)
                        if label:
                            language_id = lang.id
                            check=Translation.objects.filter(label=label,language_id=language_id)
                            if not check:
                                new_label = Translation.objects.create(label=label,language_id=language_id)
                            label_rec = Translation.objects.get(label=label,language_id=language_id)
                            trans = TranslationList.objects.filter(list_id=list_id, translation_id=label_rec.id)
                            if not trans:
                                TranslationList.objects.create(list_id=list_id, translation_id=label_rec.id)
                        if category:
                            menu_rec = Menu.objects.filter(menu_category=category,list_id=list_id)
                            if not menu_rec:
                                mdict = {}
                                mdict['menu_category']=category
                                mdict['list']=list_id
                                mrec=MenuSerializer(data=mdict, context={'request': request})
                                if mrec.is_valid():
                                    mrec.save()
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class ListFiltersViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows ListFilters to be modified.
    """
    queryset = ListFilters.objects.all()
    serializer_class = ListFilterSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class ColumnsViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Columns to be modified.
    """
    queryset = Column.objects.all()
    serializer_class = ColumnsSerializer  
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    # filterset_fields = ("app__name",)
    filterset_fields = {
            'position': ['exact'],'visibility': ['exact']
        }
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                defective_data=[]
                for i in range(len(data_dict)):
                    column = data_dict[i]['column']
                    field = data_dict[i]['field']
                    list = data_dict[i]['list']
                    visibility = data_dict[i]['visibility']
                    list_rec = List.objects.filter(system_name=list)
                    visibility_rec = Choice.objects.filter(choice_name=visibility)
                    if column and field and list_rec and visibility_rec:
                        list_id = list_rec.values()[0]['id']
                        data_dict[i]['list'] = list_id
                        columnrec = Column.objects.filter(column= column,list_id=list_id)
                        if not columnrec:
                            serializer = ColumnsSerializer(data=data_dict[i],context={'request': request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count = count+1
                    if not list_rec:
                        defective_data.append(list)
                if defective_data:
                    defective_data = {
                        "missing_lists" : f"These {defective_data} are the invalid list names."
                    } 
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
        
class MenuViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Country to be modified.
    """
    queryset = Menu.objects.all()
    serializer_class = MenuSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    # filterset_fields = ("__all__")
    filterset_fields = {
        'menu_category': ['exact', 'contains'],
        'list': ['exact'],
        'sequence': ['exact']
    }
    ordering_fields = ("__all__")
    
    def list(self, request, *args, **kwargs):
        queryset = Menu.objects.filter().all()
        serializers = MenuSerializer(queryset, many = True, context = {"request": request})
        return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                defective_data = []
                for i in range(len(data_dict)):
                    if data_dict[i] != None:
                        menu_category = data_dict[i]['menu_category']
                        if menu_category:
                            list = data_dict[i]['list']
                            sequence = data_dict[i]['sequence']
                            if list:
                                listrec = List.objects.filter(list=list)
                                if listrec:
                                    data_dict[i]['list'] = listrec.values()[0]['id']
                                    list_id = listrec.values()[0]['id']
                                    menu_rec = Menu.objects.filter(menu_category=menu_category, list_id = list_id, sequence = sequence)
                                else:
                                    defective_data.append(list)
                            else:
                                data_dict[i]['list'] = None
                                menu_rec = Menu.objects.filter(menu_category=menu_category)
                            label = data_dict[i].pop('label (us english)')
                            language = get_current_user_language(request.user)
                            lang = Language.objects.filter(name=language)
                            check=Translation.objects.filter(label=label, language_id=lang.values()[0]['id'])
                            if not check:
                                new_label = Translation.objects.create(label=label, language_id=lang.values()[0]['id'])
                            label_rec = Translation.objects.filter(label=label, language_id=lang.values()[0]['id'])
                            if not menu_rec:
                                menu_serializer = MenuSerializer(data=data_dict[i], context={'request': request})
                                if menu_serializer.is_valid(raise_exception=True):
                                    menu_serializer.save()
                                    count = count + 1
                                    new_trans = TranslationMenu.objects.create(menu_id=menu_serializer.data.get('id'), translation_id=label_rec.values()[0]['id'])
                if defective_data:
                    defective_data = {
                        "invalid_lists" : f"These {set(defective_data)} are the invalid list name(s)."
                    } 
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            print(str(e))
            return Response(utils.error(str(e)))
    
     
class HelpViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Country to be modified.
    """
    queryset = Help.objects.all()
    serializer_class = HelpSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class FormListViewSet(viewsets.ModelViewSet):
    queryset = FormList.objects.all()
    serializer_class = FormListSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class FormDataViewSet(viewsets.ModelViewSet):
    queryset = FormData.objects.all()
    serializer_class = FormDataSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                defective_data=[]
                for i in range(len(data_dict)):
                    data=data_dict[i]['data']
                    field = data_dict[i]['field']
                    if data and field:
                        form_name=data_dict[i]['form']
                        formrec = Form.objects.filter(form=form_name)
                        if formrec:
                            columnrec = FormData.objects.filter(data=data,form_id=formrec.values()[0]['id'])
                            if not columnrec:
                                data_dict[i]['form']=formrec.values()[0]['id']
                                data_dict[i]['type']=(data_dict[i]['type']).lower()
                                serializer=FormDataSerializer(data=data_dict[i], context={'request':request})
                                if serializer.is_valid(raise_exception=True):
                                    serializer.save()
                                    count += 1
                        else:
                            defective_data.append(form_name)
                if defective_data:
                    defective_data = {
                        "missing_forms" : f"These {set(defective_data)} are the invalid form names."
                    } 
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class FormSectionViewSet(viewsets.ModelViewSet):
    queryset = FormSection.objects.all()
    serializer_class = RelatedFormSectionSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class IconViewSet(viewsets.ModelViewSet):
    queryset = Icons.objects.all()
    serializer_class = IconSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("system_name",)
    ordering_fields = ("system_name",)

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                wb = openpyxl.load_workbook(file)
                sheet=wb.active
                image_loader = SheetImageLoader(sheet)
                for row in range(2,sheet.max_row+1):
                    data_dict={}
                    for column in "C":
                        cell_name = "{}{}".format(column, row)
                    image=image_loader.get(cell_name)
                    data_dict['image']=image
                    for row in sheet.iter_rows(min_row=2):
                        row_data = []
                        for cell in row:
                            row_data.append(cell.value)
                        print(row_data)
                    print(data_dict)
                msg="file received."
                return Response(utils.success(msg))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))